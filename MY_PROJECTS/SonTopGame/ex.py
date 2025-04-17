from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
import os
path = os.path.expanduser('~/Desktop/info.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "Test1"

ismlar = ['alijon','valijon','alisher','ulug\'bek','qodirbek','axrorbek']
familiya = ['arslonov','qahhorov','aliyev','mahmudov','olimov','asqarov']
year = [2001,1995,1999,2003,2000,2007]
davomat = [100,80,65,45,95,88,]

# Kataklarni birlashtirish
ws.merge_cells('A1:D1')
ws['A1'] = "Talabalarning ma'lumotlari"

# Kataklarga ma'lumot qo'shamiz
ws['A2'] = "Ism"
ws['B2'] = "Familiya"
ws['C2'] = "Tug'ilgan yil"
ws['D2'] = "Davomat"

for i in range(len(ismlar)):
    ws[f'A{3+i}'] = ismlar[i].title()
    ws[f'B{3+i}'] = familiya[i].title()
    ws[f'C{3+i}'] = year[i]
    ws[f'D{3+i}'] = davomat[i]

# row va column qo'shamiz
# ws.insert_rows(2)
# ws.insert_cols(3)

# row va column ni o'chiramiz
# ws.delete_rows(2)
# ws.delete_cols(3)

# sheet yaratamiz
wb.create_sheet('sheet 1')
ws.append(['baho','status','topik','natija'])
ws.append([5,'bad',5,'합격'])
ws.insert_rows(9)

# Font beramiz
cell = ws['A1']
cell.font = Font(bold=True,size=13,color='ff0000')
cell.alignment = Alignment(horizontal='center')
cell.fill = PatternFill(start_color='ffFFFF',end_color='FFFFFF',fill_type='darkUp')



for row in ws.iter_rows(min_row=2,max_row=2,min_col=1,max_col=4):
    for katak in row:
        katak.font = Font(bold=True,size=10,color='000000')
        katak.alignment = Alignment(horizontal='center')

for row in ws.iter_rows(min_row=3,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
    for cell in row:
        cell.font = Font(bold=True,size=9,color='000000')
        cell.alignment = Alignment(horizontal='center')

for col in ['A', 'B', 'C', 'D']:
    ws.column_dimensions[col].width = 14




  


wb.save("info.xlsx")
print("Exel fayl yaratildi: info.xlsx")


