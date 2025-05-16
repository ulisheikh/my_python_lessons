from tkinter import Tk, Label, Entry, Button , Frame
from openpyxl import Workbook,load_workbook
import os
from tkinter import messagebox


oyna = Tk()
oyna.title('StudentsInfo')
x_cor = 1200
y_cor = 800
oyna.geometry(f'{x_cor}x{y_cor}')


# Frame

frame = Frame(oyna, bg='white')
frame.place(x=0, y=0, width=600, height=600)

frame2 = Frame(oyna, bg='white')
frame2.place(x=600, y=0, width=600, height=600)

frame3 = Frame(oyna, bg='lightgray')
frame3.place(x=0, y=500, width=1200, height=300)



# Label matnlari
label_texts = [
    "Talabaning ismi:",
    "Talabaning familiyasi:",
    "Talabaning yoshi:",
    "Talabaning mamlakati:",
    "Talabaning bosqichi:",
    "Talabaning davomati:"
]


entries = []

# Label va Entry larni joylashtirish
for text in label_texts:
    sub_frame = Frame(frame, bg='white')  # har bir juftlik uchun kichik frame
    sub_frame.pack(anchor="w", pady=15)

    l = Label(sub_frame, text=text,width=25, pady=5,padx=30,anchor="w", bg='white')
    l.pack(side="left",pady=10)

    en = Entry(sub_frame, width=30)
    en.pack(side="left",pady=10)

    entries.append(en)
l2 = Label(frame2,width=50,height=17,bg='lightgray')
l2.pack(side="top",pady=20)

l_save = Label(frame3,text="File name: ")
l_save.place(x=400,y=100)



# Entry
save_en = Entry(frame3,width=50)
save_en.place(x=490,y=100)

# Functions

def save_data():
    if 
    filename = save_en.get()
    if os.path.exists(filename + ".xlsx"):
        wb = load_workbook(filename + ".xlsx")
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Students_Info"

        # Faqat yangi faylga sarlavhalar yoziladi
        ws.append(label_texts)

    row_data = [entry.get() for entry in entries]
    ws.append(row_data)

    wb.save(filename + ".xlsx")
    messagebox.showinfo("File saqlandi ")


save_but = Button(frame3,text="Save",height=1,width=5,command=save_data)
save_but.place(x=900,y=200)

oyna.mainloop()



