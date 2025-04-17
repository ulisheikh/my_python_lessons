from openpyxl import Workbook, load_workbook

# 📂 1. Excel Fayllar bilan Ishlash
wb = Workbook()                     # Yangi fayl yaratish
wb = load_workbook("file.xlsx")     # Mavjud faylni yuklash
wb.save("file.xlsx")                # Saqlash
wb.create_sheet("Sheet2")           # Yangi varaqlar
del wb["Sheet2"]                    # Varqni o'chirish

# 📄 2. Ishchi varaqlar (Worksheets)
ws = wb.active                      # Aktiv varaqqa murojaat
ws = wb["Sheet1"]                   # Nom bilan murojaat
ws.title = "NewTitle"               # Varaqqa yangi nom berish

ws["A1"] = "Hello"
value = ws["A1"].value

# 🔢 3. Kataklar bilan Ishlash
for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
    for cell in row:
        print(cell.value)

# 📊 4. Jadvallarni Kengaytirish (append)
ws.append(["Name", "Score"])       # Bir qatorda ma'lumot qo'shish
ws.append(["Ali", 90])

# 🔁 5. Ustun yoki Qator Qo‘shish / O‘chirish
ws.insert_rows(2)                  # 2-qatordan yangi qator qo‘shish
ws.delete_rows(2)                  # 2-qatordagi ma’lumotni o‘chirish

ws.insert_cols(1)                  # 1-ustundan yangi ustun
ws.delete_cols(1)                  # 1-ustunni o‘chirish

# 🧬 6. Kataklarni Birlashtirish va Ajratish
ws.merge_cells('A1:B1')            # A1 dan B1 gacha birlashtirish
ws.unmerge_cells('A1:B1')          # Ajratish

# 🎨 7. Katak Formatlash (Style)
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

cell = ws['A1']
cell.font = Font(bold=True, size=14, color="FF0000")
cell.alignment = Alignment(horizontal='center')
cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
cell.border = border

# 📐 8. Ustun / Qator kengligini sozlash
ws.column_dimensions['A'].width = 20
ws.row_dimensions[1].height = 30

# 🧠 9. Shartli formatlash (Conditional Formatting)
from openpyxl.formatting.rule import CellIsRule

ws.conditional_formatting.add('B2:B10',
    CellIsRule(operator='greaterThan', formula=['70'], fill=PatternFill(start_color='00FF00', fill_type='solid')))

# 🔍 10. Filtrlash va Avtomatik Sarlavha
ws.auto_filter.ref = "A1:B10"

# 🔗 11. Havolalar va formulalar
ws['C1'] = "=SUM(B2:B10)"           # Excel formulasi
ws['A1'].hyperlink = "https://google.com"

# 💾 12. Excel faylga oxirida saqlash
wb.save("final_result.xlsx")