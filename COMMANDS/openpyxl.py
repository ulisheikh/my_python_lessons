# openpyxl_demo.py

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# 1. 📂 Excel Fayl yaratish va ishchi varaq olish
wb = Workbook()
ws = wb.active
ws.title = "WordGameResults"

# 2. 📄 Yangi varaqlar bilan ishlash
wb.create_sheet("ExtraSheet")
# wb.remove(wb["ExtraSheet"])  # agar xohlasang o‘chirishing mumkin

# 3. 🔢 Kataklar bilan ishlash
ws["A1"] = "Name"
ws["B1"] = "Score"
ws["A2"] = "Ali"
ws["B2"] = 80

# 4. 📊 Jadvallarni kengaytirish
ws.append(["Vali", 90])
ws.append(["Hasan", 75])

# 5. 🔁 Ustun/qator qo‘shish va o‘chirish
ws.insert_rows(2)     # 2-qatordan yangi qator qo‘shiladi
ws.delete_rows(2)     # 2-qatordagi ma’lumotni o‘chiramiz

ws.insert_cols(3)     # C ustuni (3-index) qo‘shiladi
ws.delete_cols(3)     # C ustuni o‘chiriladi

# 6. 🧬 Kataklarni birlashtirish va ajratish
ws["D1"] = "Birlashtirilgan"
ws.merge_cells("D1:E1")          # D1 va E1 birlashtirildi
# ws.unmerge_cells("D1:E1")      # Ajratish uchun

# 7. 🎨 Formatlash: shrift, markazlash, rang va chiziq
cell = ws["A1"]
cell.font = Font(bold=True, size=14, color="0000FF")
cell.alignment = Alignment(horizontal="center")
cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
cell.border = border

# 8. 📐 Ustun va qator o‘lchamini sozlash
ws.column_dimensions["A"].width = 20
ws.row_dimensions[1].height = 25

# 9. 🧠 Shartli formatlash – agar qiymat 70 dan katta bo‘lsa yashil bo‘lsin
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
ws.conditional_formatting.add("B2:B10", CellIsRule(operator='greaterThan', formula=['70'], fill=green_fill))

# 10. 🔍 Avtofiltr qo‘shish
ws.auto_filter.ref = "A1:B10"

# 11. 🔗 Formulalar va havolalar
ws["C1"] = "=SUM(B2:B10)"            # Summasi
ws["A5"] = "Visit Google"
ws["A5"].hyperlink = "https://google.com"

# 12. 💾 Oxirida faylni saqlash
wb.save("openpyxl_demo.xlsx")

print("✅ Excel fayl yaratildi: openpyxl_demo.xlsx")
